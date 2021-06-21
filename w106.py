#! /usr/bin/env python
# sudo pip install pyModbusTCP
from pyModbusTCP.client import ModbusClient
import datetime
from influxdb import InfluxDBClient
from env import *
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path

if ((len(sys.argv)) == 3):
    mode = sys.argv[2]
    if (sys.argv[1] == "1"):
        ipAddress = "192.168.88.10"
        port = 8080
        dbname = "H1V1A01"
    if (sys.argv[1] == "2"):
        ipAddress = "192.168.88.11"
        port = 8080
        dbname = "F1V1A01"
    if (sys.argv[1] == "3"):
        ipAddress = "192.168.88.12"
        port = 8080
        dbname = "A1V1A02"
    if (sys.argv[1] == "4"):
        ipAddress = "192.168.88.13"
        port = 8080
        dbname = "A1V1A01"
    if (sys.argv[1] == "5"):
        ipAddress = "192.168.88.14"
        port = 8080
        dbname = "D1V1A01"
    if (sys.argv[1] == "6"):
        ipAddress = "192.168.88.15"
        port = 8080
        dbname = "MainLab1"
    if (sys.argv[1] == "7"):
        ipAddress = "192.168.88.16"
        port = 8080
        dbname = "MainLab2"
    if (sys.argv[1] == "8"):
        ipAddress = "192.168.88.30"
        port = 8080
        dbname = "Diesel"
    if (sys.argv[1] == "9"):
        ipAddress = "192.168.88.31"
        port = 8080
        dbname = "B1"
    if (sys.argv[1] == "10"):
        ipAddress = "192.168.88.32"
        port = 8080
        dbname = "B2"
    if (sys.argv[1] == "11"):
        ipAddress = "192.168.88.33"
        port = 8080
        dbname = "B3"

    if (sys.argv[1] == "12"):
        ipAddress = "192.168.88.34"
        port = 8080
        dbname = "T1"
    if (sys.argv[1] == "13"):
        ipAddress = "192.168.88.35"
        port = 8080
        dbname = "T2"
    if (sys.argv[1] == "14"):
        ipAddress = "192.168.88.36"
        port = 8080
        dbname = "T3"                
    if (sys.argv[1] == "15"):
        ipAddress = "192.168.88.37"
        port = 8080
        dbname = "LMDP"       
    if (sys.argv[1] == "16"):
        ipAddress = "192.168.88.38"
        port = 8080
        dbname = "ELMDP"       
    if (sys.argv[1] == "17"):
        ipAddress = "192.168.88.39"
        port = 8080
        dbname = "sole_sefid"
    if (sys.argv[1] == "18"):
        ipAddress = "192.168.88.28"
        port = 8080
        dbname = "Ice_bank"
    if (sys.argv[1] == "19"):
        ipAddress = "192.168.88.25"
        port = 8080
        dbname = "Comp1"
    if (sys.argv[1] == "20"):
        ipAddress = "192.168.88.26"
        port = 8080
        dbname = "Comp2"        
    if (sys.argv[1] == "21"):
        ipAddress = "192.168.88.27"
        port = 8080
        dbname = "Comp3" 


elif ((len(sys.argv)) == 5):
    ipAddress = sys.argv[1]
    port = sys.argv[2]
    dbname = sys.argv[3]
    mode = sys.argv[4]
else:
    print('less arguments!')
    print('STOPED!')
    exit()



time = datetime.datetime.utcnow()


def excelWrite(dbname, CounterA1, CounterA1peak, CounterA1normal, CounterA1low, CounterP1, CounterA2, CounterP2):
    myFileName = os.path.dirname(os.path.abspath(__file__))+"/counter/"+ dbname +".xlsx"
    if not (os.path.isfile(myFileName)):
        wb = Workbook()
        #wb.save(myFileName)
        #wb = load_workbook(filename=myFileName)
        ws = wb['Sheet']
        ws.cell(column=1, row=1, value=dbname)
        ws.cell(column=2, row=1, value="A+")
        ws.cell(column=3, row=1, value="A+peak")
        ws.cell(column=4, row=1, value="A+normal")
        ws.cell(column=5, row=1, value="A+low")
        ws.cell(column=6, row=1, value="P+")
        ws.cell(column=7, row=1, value="A-")
        ws.cell(column=8, row=1, value="P-")
        wb.save(myFileName)
    
    timenow = datetime.datetime.now()
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet']
    newRowLocation = ws.max_row + 1
    ws.cell(column=1, row=newRowLocation, value=timenow)
    ws.cell(column=2, row=newRowLocation, value=CounterA1)
    ws.cell(column=3, row=newRowLocation, value=CounterA1peak)
    ws.cell(column=4, row=newRowLocation, value=CounterA1normal)
    ws.cell(column=5, row=newRowLocation, value=CounterA1low)
    ws.cell(column=6, row=newRowLocation, value=CounterP1)
    ws.cell(column=7, row=newRowLocation, value=CounterA2)
    ws.cell(column=8, row=newRowLocation, value=CounterP2)
    wb.save(filename=myFileName)
    wb.close()




class W106:

    def __init__(self, address, port, mode):
        c = ModbusClient()
        c.host(address)
        c.port(port)
        c.unit_id(1)
        c.open()
        if (mode == "volt"):
            voltAmper = c.read_input_registers(0, 50)
            c.close()
            if voltAmper:
                self.v1 = (voltAmper[0] << 16 | voltAmper[1])/10
                self.v2 = (voltAmper[2] << 16 | voltAmper[3])/10
                self.v3 = (voltAmper[4] << 16 | voltAmper[5])/10
                self.vavg = (voltAmper[6] << 16 | voltAmper[7])/10
                self.vun = (voltAmper[8] << 16 | voltAmper[9])/10

                self.v12 = (voltAmper[10] << 16 | voltAmper[11])/10
                self.v23 = (voltAmper[12] << 16 | voltAmper[13])/10
                self.v31 = (voltAmper[14] << 16 | voltAmper[15])/10

                self.i1 = (voltAmper[16] << 16 | voltAmper[17])/10
                self.i2 = (voltAmper[18] << 16 | voltAmper[19])/10
                self.i3 = (voltAmper[20] << 16 | voltAmper[21])/10
                self.iavg = (voltAmper[22] << 16 | voltAmper[23])/10
                self.inut = (voltAmper[24] << 16 | voltAmper[25])/10

                self.ptot = (voltAmper[44] << 16 | voltAmper[45])/10
                self.qtot = (voltAmper[46] << 16 | voltAmper[47])/10
                self.stot = (voltAmper[48] << 16 | voltAmper[49])/10

            else:
                print("Read Volt And Amper ERROR")

        if (mode == "counter"):
            Counter = c.read_input_registers(70, 36)
            #print(Counter)
            c.close()
            if Counter:
                self.peakA1 = (Counter[0] << 32 |
                               Counter[1] << 16 | Counter[2])
                self.peakP1 = (Counter[3] << 32 |
                               Counter[4] << 16 | Counter[5])
                self.peakA2 = (Counter[6] << 32 |
                               Counter[7] << 16 | Counter[8])
                self.peakP2 = (Counter[9] << 32 |
                               Counter[10] << 16 | Counter[11])

                self.normalA1 = (Counter[12] << 32 |
                                 Counter[13] << 16 | Counter[14])
                self.normalP1 = (Counter[15] << 32 |
                                 Counter[16] << 16 | Counter[17])
                self.normalA2 = (Counter[18] << 32 |
                                 Counter[19] << 16 | Counter[20])
                self.normalP2 = (Counter[21] << 32 |
                                 Counter[22] << 16 | Counter[23])

                self.lowA1 = (Counter[24] << 32 |
                              Counter[25] << 16 | Counter[26])
                self.lowP1 = (Counter[27] << 32 |
                              Counter[28] << 16 | Counter[29])
                self.lowA2 = (Counter[30] << 32 |
                              Counter[31] << 16 | Counter[32])
                self.lowP2 = (Counter[33] << 32 |
                              Counter[34] << 16 | Counter[35])

                self.CounterA1 = self.peakA1 + self.normalA1 + self.lowA1
                self.CounterP1 = self.peakP1 + self.normalP1 + self.lowP1
                self.CounterA2 = self.peakA2 + self.normalA2 + self.lowA2
                self.CounterP2 = self.peakP2 + self.normalP2 + self.lowP2

            else:
                print("Read Counter ERROR")

w106 = W106(ipAddress, port, mode)
if (mode == "volt"):
    measurement_name = dbname
    body = [
        {
            "measurement": measurement_name,
            "time": time,
            "fields": {
                "V1": w106.v1,
                "V2": w106.v2,
                "V3": w106.v3,
                "Vavg": w106.vavg,
                "I1": w106.i1,
                "I2": w106.i2,
                "I3": w106.i3,
                "Iavg": w106.iavg,
                "Ptot": w106.ptot,
                "Qtot": w106.qtot,
                "Stot": w106.stot,
            }
        }
    ]

if (mode == "counter"):
    measurement_name = dbname + "-counter"
    excelWrite(dbname,w106.CounterA1, w106.peakA1, w106.normalA1, w106.lowA1, w106.CounterP1, w106.CounterA2, w106.CounterP2)
    body = [
        {
            "measurement": measurement_name,
            "time": time,
            "fields": {
                "A+": w106.CounterA1,
                "A+peak": w106.peakA1,
                "A+normal": w106.normalA1,
                "A+low": w106.lowA1,
                "P+": w106.CounterP1,
                "A-": w106.CounterA2,
                "P-": w106.CounterP2,
            }
        }
    ]

ifclient = InfluxDBClient(ifhost, ifport, ifuser, ifpass, ifdb)
ifclient.write_points(body)
